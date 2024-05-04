# -*- coding: utf-8 -*-
"""
Created on Sat May  4 18:26:01 2024

@author: prana
"""



import tkinter as tk
from tkinter import messagebox
import openpyxl
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import os

# Function to add expenses
def add_expenses(user_name, income, month):
    # Declare global variables for entry fields
    global grocery_budget_entry, clothing_budget_entry, travelling_budget_entry
    global grocery_expense_entry, clothing_expense_entry, travelling_expense_entry, other_expenses_entry, goal_entry
    global savings_label  # Define savings_label as global
    
    # Function to submit expenses
    def submit_expenses(user_name, income, month):  
        # Retrieve values from entry fields and convert them to integers
        grocery_budget = int(grocery_budget_entry.get())
        clothing_budget = int(clothing_budget_entry.get())
        travelling_budget = int(travelling_budget_entry.get())
        grocery_expense = int(grocery_expense_entry.get())
        clothing_expense = int(clothing_expense_entry.get())
        travelling_expense = int(travelling_expense_entry.get())
        other_expenses = int(other_expenses_entry.get())
        goal = int(goal_entry.get())

        # Create a list of expenses
        expenses = [[grocery_budget, grocery_expense], [clothing_budget, clothing_expense], [travelling_budget, travelling_expense], other_expenses, goal]
        
        # Calculate savings
        grocery_savings, clothing_savings, travelling_savings, other_savings, total_savings = calculate_savings(expenses, income)
        
        # Update the savings label
        savings_label.config(text=f"Total Savings: {total_savings}")
        
        # Generate expense report
        generate_report(user_name, income, expenses, total_savings, month, filename)

    # Create a new window for entering expenses
    expenses_window = tk.Toplevel()
    expenses_window.title("Enter Expenses")
    
    # Labels and Entry fields for expenses
    tk.Label(expenses_window, text="Enter budget for grocery:").grid(row=0, column=0)
    grocery_budget_entry = tk.Entry(expenses_window)
    grocery_budget_entry.grid(row=0, column=1)

    tk.Label(expenses_window, text="Enter budget for clothing:").grid(row=1, column=0)
    clothing_budget_entry = tk.Entry(expenses_window)
    clothing_budget_entry.grid(row=1, column=1)

    tk.Label(expenses_window, text="Enter budget for travelling:").grid(row=2, column=0)
    travelling_budget_entry = tk.Entry(expenses_window)
    travelling_budget_entry.grid(row=2, column=1)

    tk.Label(expenses_window, text="Enter actual expense on grocery:").grid(row=3, column=0)
    grocery_expense_entry = tk.Entry(expenses_window)
    grocery_expense_entry.grid(row=3, column=1)

    tk.Label(expenses_window, text="Enter actual expense on clothing:").grid(row=4, column=0)
    clothing_expense_entry = tk.Entry(expenses_window)
    clothing_expense_entry.grid(row=4, column=1)

    tk.Label(expenses_window, text="Enter actual expense on travelling:").grid(row=5, column=0)
    travelling_expense_entry = tk.Entry(expenses_window)
    travelling_expense_entry.grid(row=5, column=1)

    tk.Label(expenses_window, text="Enter your other expenses:").grid(row=6, column=0)
    other_expenses_entry = tk.Entry(expenses_window)
    other_expenses_entry.grid(row=6, column=1)

    tk.Label(expenses_window, text="Enter your overall budget goal:").grid(row=7, column=0)
    goal_entry = tk.Entry(expenses_window)
    goal_entry.grid(row=7, column=1)

    # Button to submit expenses
    submit_button = tk.Button(expenses_window, text="Submit", command=lambda: submit_expenses(user_name, income, month))
    submit_button.grid(row=8, column=0, columnspan=2)

# Function to calculate savings
def calculate_savings(expenses, income):
    grocery_savings = expenses[0][0] - expenses[0][1]
    clothing_savings = expenses[1][0] - expenses[1][1]
    travelling_savings = expenses[2][0] - expenses[2][1]
    other_savings = income - expenses[3] - sum([expense[1] for expense in expenses[:3]])
    total_savings = grocery_savings + clothing_savings + travelling_savings + other_savings
    return grocery_savings, clothing_savings, travelling_savings, other_savings, total_savings

# Function to generate expense report
def generate_report(user_name, income, expenses, total_savings, month, filename):
    global report_canvas
    
    # Create a new window for expense report
    report_window = tk.Toplevel()
    report_window.title("Expense Report")
    
    # Display user summary
    user_summary_label = tk.Label(report_window, text=f'User: {user_name}')
    user_summary_label.pack()

    # Display expense categories
    tk.Label(report_window, text='Expense Category').pack()
    categories = ['Grocery', 'Clothing', 'Travelling', 'Other']
    for category in categories:
        if isinstance(expenses[categories.index(category)], list):
            budget = expenses[categories.index(category)][0]
            expense = expenses[categories.index(category)][1]
            savings = budget - expense
        else:
            budget = 'N/A'
            expense = expenses[categories.index(category)]
            savings = 'N/A'
        tk.Label(report_window, text=f'{category}: Budget={budget}, Expense={expense}, Savings={savings}').pack()

    # Display total savings
    total_savings_label = tk.Label(report_window, text=f'Total Savings: {total_savings}')
    total_savings_label.pack()

    # Display goal achievement message
    goal_achievement_label = tk.Label(report_window, text='Congratulations! You have achieved your overall budget goal.' if total_savings >= expenses[4] else 'You have not achieved your overall budget goal.')
    goal_achievement_label.pack()

    # Open or create Excel workbook
    if os.path.isfile(filename):
        workbook = openpyxl.load_workbook(filename)
        worksheet = workbook.active
    else:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = 'Expense Report'
        headers = ['User', 'Total Savings', 'Goal Achievement', 'Month', 'Grocery Actual Expense', 'Clothing Actual Expense', 'Travelling Actual Expense', 'Other Actual Expense']
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)

    # Write data to Excel workbook
    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=user_name)
    worksheet.cell(row=row, column=2, value=total_savings)
    worksheet.cell(row=row, column=3, value='Achieved' if total_savings >= expenses[4] else 'Not Achieved')
    worksheet.cell(row=row, column=4, value=month)
    
    for i, category in enumerate(categories):
        if isinstance(expenses[i], list):
            worksheet.cell(row=row, column=i+5, value=expenses[i][1])
        else:
            worksheet.cell(row=row, column=i+5, value=expenses[i])

    workbook.save(filename)
    # Show message box with file save information
    messagebox.showinfo("Expense Report", f"Expense report saved to {filename}")

    # Generate and display expense graphs
    generate_graphs(report_window, expenses)

# Function to generate expense comparison and distribution graphs
def generate_graphs(report_window, expenses):
    categories = ['Grocery', 'Clothing', 'Travelling', 'Other']
    filtered_expenses = [expense for expense in expenses if isinstance(expense, list)]
    filtered_labels = [category for expense, category in zip(filtered_expenses, categories) if isinstance(expense, list)]
    
    # Create a frame to contain the charts
    charts_frame = tk.Frame(report_window)
    charts_frame.pack()

    # Create a canvas for the bar chart
    bar_chart_canvas = FigureCanvasTkAgg(plt.figure(figsize=(5, 4)), master=charts_frame)
    bar_chart_canvas.get_tk_widget().pack(side=tk.LEFT, fill=tk.BOTH, expand=1)
    
    # Generate the bar chart
    expenses_data = [expense[1] for expense in filtered_expenses] + [expenses[categories.index('Other')]]
    plt.bar(categories, expenses_data)
    plt.title('Expense Comparison')
    plt.xlabel('Expense Category')
    plt.ylabel('Expense Amount')
    bar_chart_canvas.draw()
    
    # Create a canvas for the pie chart
    pie_chart_canvas = FigureCanvasTkAgg(plt.figure(figsize=(5, 4)), master=charts_frame)
    pie_chart_canvas.get_tk_widget().pack(side=tk.RIGHT, fill=tk.BOTH, expand=1)
    
    # Generate the pie chart
    if filtered_expenses:
        plt.pie([expense[1] for expense in filtered_expenses], labels=filtered_labels, autopct='%1.1f%%', startangle=140)
        plt.title('Expense Distribution')
    else:
        plt.text(0.5, 0.5, 'No data available', horizontalalignment='center', verticalalignment='center')
    pie_chart_canvas.draw()

# Create the main Tkinter window
root = tk.Tk()
root.title("Expense Tracker")

# Labels and Entry fields for user input
user_name_label = tk.Label(root, text="Enter your name:")
user_name_label.pack()
user_name_entry = tk.Entry(root)
user_name_entry.pack()

num_sources_label = tk.Label(root, text="Enter the number of your income sources:")
num_sources_label.pack()
num_sources_entry = tk.Entry(root)
num_sources_entry.pack()

income_label = tk.Label(root, text="Enter your total income through all sources for 1 month:")
income_label.pack()
income_entry = tk.Entry(root)
income_entry.pack()

month_label = tk.Label(root, text="Enter the month:")
month_label.pack()
month_entry = tk.Entry(root)
month_entry.pack()

filename = "C:\\Users\\prana\\OneDrive\\Dokumen\\PROJECT_FINAL.xlsx"

savings_label = tk.Label(root, text="")
savings_label.pack()

# Button to input expenses
input_expenses_button = tk.Button(root, text="Input Expenses", command=lambda: add_expenses(user_name_entry.get(), int(income_entry.get()), month_entry.get()))
input_expenses_button.pack()

# Start the main event loop
root.mainloop()
